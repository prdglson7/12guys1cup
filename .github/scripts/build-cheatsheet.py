"""
Builds 12guys1cup-cheat-sheet.xlsx from tiers-adp.json (FFBallers tiers + ADP).
Runs after fetch-draftkit.js in the Draft Kit workflow.

Layout matches FFBallers spreadsheet:
  - Sheet 1: 4 positions side-by-side (QB | RB | WR | TE) with tier breaks
  - Sheet 2: Top 200 overall (excludes K & DST)
  - Columns: Rank, Player (Team), BYE, ADP
"""
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
TIERS_FILE = os.path.join(REPO_ROOT, 'assets', 'data', 'tiers-adp.json')
OUTPUT = os.path.join(REPO_ROOT, 'assets', 'data', '12guys1cup-cheat-sheet.xlsx')

if not os.path.exists(TIERS_FILE):
    print(f'tiers-adp.json not found at {TIERS_FILE} — skipping cheat sheet build.')
    exit(0)

with open(TIERS_FILE) as f:
    tiers_raw = json.load(f)

# Build per-position lists (preserving parse order = FFBallers rank order)
positions = ['QB', 'RB', 'WR', 'TE']
by_pos = {p: [] for p in positions}
for key, val in tiers_raw.items():
    pos = val.get('pos')
    if pos in by_pos:
        by_pos[pos].append(val)

# ─── Styles ───
FNT = "Calibri"
navy = "1E3A5F"
gold = "E8B84A"
green = "4CAF50"

title_font = Font(name=FNT, bold=True, size=14, color=gold)
title_fill = PatternFill("solid", fgColor=navy)
sub_font = Font(name=FNT, size=10, color="FFFFFF")
hdr_font = Font(name=FNT, bold=True, size=10, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="000000")
hdr_align = Alignment(horizontal="center", vertical="center")
tier_font = Font(name=FNT, bold=True, size=10, color="FFFFFF")
tier_fill = PatternFill("solid", fgColor=green)
tier_align = Alignment(horizontal="left", vertical="center")
data_font = Font(name=FNT, size=11)
align_c = Alignment(horizontal="center", vertical="center")
align_l = Alignment(horizontal="left", vertical="center")
thin = Side(style="thin", color="E0E0E0")
cell_border = Border(bottom=thin)

pos_colors = {'QB': 'B8386B', 'RB': '2E7D32', 'WR': '1565C0', 'TE': 'E65100'}

wb = openpyxl.Workbook()

# ════════════ SHEET 1: Cheat Sheet ════════════
ws = wb.active
ws.title = "Cheat Sheet"
ws.sheet_properties.tabColor = navy

# Title rows
ws.merge_cells("A1:S1")
ws["A1"].value = "12 GUYS 1 CUP — DRAFT CHEAT SHEET"
ws["A1"].font = title_font
ws["A1"].fill = title_fill
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:S2")
ws["A2"].value = "Full PPR · 1QB/2RB/2WR/1TE/1FLEX(RB/WR/TE) · 12 Teams · Draft Day: Sept 6, 2026"
ws["A2"].font = sub_font
ws["A2"].fill = title_fill
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 6

# Position blocks: QB=A-D, gap=E, RB=F-I, gap=J, WR=K-N, gap=O, TE=P-S
blocks = [
    {'pos': 'QB', 'start': 1},
    {'pos': 'RB', 'start': 6},
    {'pos': 'WR', 'start': 11},
    {'pos': 'TE', 'start': 16},
]

# Position headers (row 4)
for block in blocks:
    pos, s = block['pos'], block['start']
    ws.merge_cells(start_row=4, start_column=s, end_row=4, end_column=s+3)
    cell = ws.cell(row=4, column=s)
    cell.value = {'QB':'Quarterbacks','RB':'Running Backs','WR':'Wide Receivers','TE':'Tight Ends'}[pos]
    cell.font = Font(name=FNT, bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=pos_colors[pos])
    cell.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(s, s+4):
        ws.cell(row=4, column=c).fill = PatternFill("solid", fgColor=pos_colors[pos])
ws.row_dimensions[4].height = 24

# Column headers (row 5)
for block in blocks:
    s = block['start']
    for i, h in enumerate(['#', 'Player', 'BYE', 'ADP']):
        cell = ws.cell(row=5, column=s+i, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align

# Column widths
for block in blocks:
    s = block['start']
    ws.column_dimensions[get_column_letter(s)].width = 4
    ws.column_dimensions[get_column_letter(s+1)].width = 28
    ws.column_dimensions[get_column_letter(s+2)].width = 5
    ws.column_dimensions[get_column_letter(s+3)].width = 6
for gap in [5, 10, 15]:
    ws.column_dimensions[get_column_letter(gap)].width = 2

# Player data with tier breaks
pos_indices = {b['pos']: 0 for b in blocks}
tier_tracker = {b['pos']: None for b in blocks}
row = 6

while any(pos_indices[b['pos']] < len(by_pos[b['pos']]) for b in blocks):
    # Check for tier breaks
    need_tier = False
    for block in blocks:
        pos = block['pos']
        idx = pos_indices[pos]
        if idx < len(by_pos[pos]):
            t = by_pos[pos][idx].get('tier')
            if t and t != tier_tracker[pos]:
                need_tier = True

    if need_tier:
        for block in blocks:
            pos, s = block['pos'], block['start']
            idx = pos_indices[pos]
            if idx < len(by_pos[pos]):
                t = by_pos[pos][idx].get('tier')
                if t and t != tier_tracker[pos]:
                    tier_tracker[pos] = t
                    ws.merge_cells(start_row=row, start_column=s, end_row=row, end_column=s+3)
                    cell = ws.cell(row=row, column=s)
                    cell.value = f"TIER {t}"
                    cell.font = tier_font
                    cell.fill = tier_fill
                    cell.alignment = tier_align
                    for c in range(s, s+4):
                        ws.cell(row=row, column=c).fill = tier_fill
        row += 1

    # Write player row
    for block in blocks:
        pos, s = block['pos'], block['start']
        idx = pos_indices[pos]
        if idx < len(by_pos[pos]):
            p = by_pos[pos][idx]
            name_team = f"{p['name']} ({p['team']})" if p.get('team') else p['name']
            adp = p.get('adp', '')
            bye = p.get('bye', '')

            for ci in range(4):
                ws.cell(row=row, column=s+ci).border = cell_border

            ws.cell(row=row, column=s, value=idx+1).font = data_font
            ws.cell(row=row, column=s).alignment = align_c
            ws.cell(row=row, column=s+1, value=name_team).font = data_font
            ws.cell(row=row, column=s+1).alignment = align_l

            try: bye_val = int(bye) if bye else None
            except: bye_val = bye or None
            ws.cell(row=row, column=s+2, value=bye_val).font = data_font
            ws.cell(row=row, column=s+2).alignment = align_c
            ws.cell(row=row, column=s+3, value=adp if adp else '—').font = data_font
            ws.cell(row=row, column=s+3).alignment = align_c

            pos_indices[pos] += 1
    row += 1

ws.freeze_panes = "A6"

# ════════════ SHEET 2: Top 200 ════════════
ws2 = wb.create_sheet("Top 200")
ws2.sheet_properties.tabColor = navy

# Build overall ranking from the LG# field — players with no LG# go to end
all_players = []
for pos in ['QB', 'RB', 'WR', 'TE']:
    for p in by_pos[pos]:
        all_players.append(p)

# Sort by position rank within FFBallers data (parse order approximates this)
# Use a simple interleave: rank all players by their FFBallers-assigned LG rank if available
# Since tiers-adp.json doesn't have LG#, we'll interleave by tier + position order
# Approximation: sort by tier first, then by position parse order
all_players.sort(key=lambda p: (p.get('tier', 99), positions.index(p['pos']) if p['pos'] in positions else 99))

top200 = all_players[:200]

ws2.merge_cells("A1:E1")
ws2["A1"].value = "12 GUYS 1 CUP — TOP 200 OVERALL"
ws2["A1"].font = title_font
ws2["A1"].fill = title_fill
ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[1].height = 28

ws2.merge_cells("A2:E2")
ws2["A2"].value = "Excludes K & DST · Full PPR · 1QB/2RB/2WR/1TE/1FLEX · 12 Teams"
ws2["A2"].font = sub_font
ws2["A2"].fill = title_fill
ws2.row_dimensions[2].height = 18

for i, h in enumerate(['#', 'Player', 'Pos', 'BYE', 'ADP'], 1):
    cell = ws2.cell(row=3, column=i, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align

ws2.column_dimensions['A'].width = 4
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 5
ws2.column_dimensions['D'].width = 5
ws2.column_dimensions['E'].width = 6

pos_fill_map = {
    'QB': PatternFill("solid", fgColor="F8D7E8"),
    'RB': PatternFill("solid", fgColor="D5ECD5"),
    'WR': PatternFill("solid", fgColor="D4E4F7"),
    'TE': PatternFill("solid", fgColor="FCE4D6"),
}
pos_font_map = {
    'QB': Font(name=FNT, size=10, bold=True, color="B8386B"),
    'RB': Font(name=FNT, size=10, bold=True, color="2E7D32"),
    'WR': Font(name=FNT, size=10, bold=True, color="1565C0"),
    'TE': Font(name=FNT, size=10, bold=True, color="E65100"),
}

for i, p in enumerate(top200):
    r = i + 4
    name_team = f"{p['name']} ({p['team']})" if p.get('team') else p['name']

    for ci in range(1, 6):
        ws2.cell(row=r, column=ci).border = cell_border

    ws2.cell(row=r, column=1, value=i+1).font = data_font
    ws2.cell(row=r, column=1).alignment = align_c
    ws2.cell(row=r, column=2, value=name_team).font = data_font
    ws2.cell(row=r, column=2).alignment = align_l

    pos_cell = ws2.cell(row=r, column=3, value=p['pos'])
    pos_cell.font = pos_font_map.get(p['pos'], data_font)
    pos_cell.fill = pos_fill_map.get(p['pos'], PatternFill())
    pos_cell.alignment = align_c

    try: bye_v = int(p['bye']) if p.get('bye') else None
    except: bye_v = p.get('bye') or None
    ws2.cell(row=r, column=4, value=bye_v).font = data_font
    ws2.cell(row=r, column=4).alignment = align_c
    ws2.cell(row=r, column=5, value=p.get('adp','') or '—').font = data_font
    ws2.cell(row=r, column=5).alignment = align_c

ws2.freeze_panes = "A4"

wb.save(OUTPUT)
print(f'Built {OUTPUT}')
print(f'  Sheet 1: {sum(len(by_pos[p]) for p in positions)} players')
print(f'  Sheet 2: {len(top200)} players')
